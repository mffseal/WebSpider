import requests
import bs4
import openpyxl
import re
import threading
import math


# 抓取整个页面
def get_url(url, use_proxy=True, wait_time=5):
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:62.0) Gecko/20100101 Firefox/62.0'}
    # 代理设置分别匹配http和https协议
    # 要支持socks5要先pip install -U requests[socks]
    proxies = {"http": "socks5://127.0.0.1:1080", "https": "socks5://127.0.0.1:1080"}
    if use_proxy is True:
        res = requests.get(url, headers=headers, proxies=proxies, timeout=wait_time)
    else:
        res = requests.get(url, headers=headers, timeout=wait_time)

    return res


# 获取页数

# 获取商品信息
def get_info(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    photo_url = []
    item_url = []
    title = []
    identify = []

    # 定位商品标题
    title_pond = soup.find_all('a', class_='vip')
    # 提取标题文字和商品详情链接
    for each in title_pond:
        identify.append(filter_title(each.text))
        title.append(each.text)  # 标题
        item_url.append(each.get("href"))  # 链接

    # 提取图片链接
    img_pond = soup.find_all('img', class_='img')
    for each in img_pond:
        photo_url.append(each.get("src"))

    # 导出信息
    info = [title, photo_url, identify, item_url]
    return info


# 获取图片链接
def get_img_url(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    photo_url = []

    img_pond = soup.find_all('img', class_='img')
    for each in img_pond:
        photo_url.append(each.get("src"))

    return photo_url


# 获取标题
def get_title(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    title = []

    # 定位商品标题
    title_pond = soup.find_all('a', class_='vip')
    # 提取标题文字和商品详情链接
    for each in title_pond:
        title.append(each.text)  # 标题

    return title


# 获取商品详情链接
def get_detail_url(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    item_url = []
    # 定位商品标题
    title_pond = soup.find_all('a', class_='vip')
    # 提取标题文字和商品详情链接
    for each in title_pond:
        item_url.append(each.get("href"))  # 链接

    return item_url


# 下载图片
def get_pic(url, count, try_time=1):
    file_name = "./img/" + str(count) + ".png"

    try:
        pic = get_url(url, False, 10)
    except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectionError):
        print("!!!第%d张图片下载失败%d次" % (count, try_time))
        # 错误3次写入错误日志
        if try_time >= 3:
            print("!!!图片下载失败3次, 写入错误日志")
            err_log = open("./img/err_log.txt", 'r+')
            err_log.write(str(count) + "\n")
            err_log.close()
        else:
            try_time += 1
            get_pic(url, count, try_time)
    else:
        open(file_name, 'wb').write(pic.content)
        print("第%d张图片下载完成" % count)


# 处理标题得到款式编号
def filter_title(title):
    # 正则, 大于5位小于20位的数字, 结尾可能有大小写字符, 可能结尾字母后还有数字, 结尾不能有标点符号
    try:
        identity = re.search(r"\d{5,20}[a-z]*[A-Z]*\d*[^.,]", title).group(0)
    except AttributeError:
        identity = "编号获取失败"
    return identity  # 排除第一个空格


# 导出到excel
def data_export(data):
    wb = openpyxl.Workbook()  # 坑爹, Workbook() 的 W 要大写
    wb.guess_types = True  # 自动匹配数据类型
    ws = wb.active
    for count in range(len(data[0])):
        d1_identify = ws.cell(row=count+1, column=1)
        d2_title = ws.cell(row=count+1, column=2)
        d3_url = ws.cell(row=count+1, column=3)
        d1_identify.value = data[0][count]
        d2_title.value = data[1][count]
        d3_url.value = data[2][count]

    wb.save('款式表.xlsx')

    # 图片导入excel: <table><img width="140" height="140" src="D:\codes\python\WebSpider\EbaySpider\img\1.png">


def get_some_pics(count_img, img_pack):
    for each in img_pack:
        count_img += 1
        get_pic(each, count_img)


# 主函数
def main():
    """
    商店列表:
    kimskouture +
    rocknation +
    fabsilvercharms +
    allforloveboutique +
    mrskendall11 +
    justsayjulie2709 +
    loriesteven
    sheilamcara1968
    brownsbug707
    loriesteven
    atomicspook
    """
    shop_name = "kimskouture"
    page_amount = 25
    total_title = []
    total_identify = []
    total_detail_url = []
    total_img = []

    # 重置图片下载错误日志
    open("./img/err_log.txt", 'w').close()

    # 获取分页信息
    for i in range(1, page_amount+1):
        print("-----开始解析第%d页-----" % i)
        url = "https://www.ebay.com/sch/m.html?_ssn=" + shop_name + "&_pgn=" + str(i)
        per_page = get_url(url)
        total_title.extend(get_title(per_page))
        print("更新: 当前收录%d个商品" % len(total_title))
        total_detail_url.extend(get_detail_url(per_page))
        total_img.extend(get_img_url(per_page))
        print("-----第%d页解析完成-----" % i)

    # 处理标题得到型号
    print("-----开始处理标题-----")
    for each in total_title:
        total_identify.append(filter_title(each))
    print("-----标题处理完毕-----")

    # 下载图片
    print("-----开始下载图片-----")
    img_amount = len(total_img)
    # 多线程
    thread_list = []
    img_pack_size = 20  # 每个进程下载几个图片
    offset = 0
    if img_amount < img_pack_size:
        thread_amount = 1
    else:
        thread_amount = math.ceil(img_amount / img_pack_size)  # 向上取整
    for i in range(thread_amount-1):  # 少创建一个进程, 最后一个进程单独考虑
        offset = i * img_pack_size
        # 创建进程list
        thread_list.append(threading.Thread(target=get_some_pics, args=(offset, total_img[offset:offset + img_pack_size]
                                                                        )))
    # 考虑最后一个线程的图片数不确定
    thread_list.append(threading.Thread(target=get_some_pics, args=(offset, total_img[offset:img_amount])))
    print("---将创建%d个下载线程---" % thread_amount)
    thread_surplus = 0
    for each in thread_list:
        each.start()  # 启动进程
        thread_surplus += 1
        print("子线程启动, 剩余线程: %d" % thread_surplus)
    for each in thread_list:
        each.join()  # 让父进程等待
        thread_surplus -= 1
        print("子线程关闭, 剩余线程: %d" % thread_surplus)
    print("-----全部图片下载完毕-----")

    # 整合数据
    print("-----开始整合数据-----")
    total_info = [total_identify, total_title, total_detail_url]
    print("-----数据整合完毕-----")

    # 执行导出
    print("-----开始导出到excel-----")
    data_export(total_info)
    print("-----数据导出完毕-----")


if __name__ == '__main__':
    # 开始爬取内容
    main()
