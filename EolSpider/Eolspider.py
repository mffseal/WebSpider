# coding=utf-8
import requests
import bs4
import openpyxl
import re


# 抓取整个页面
def get_url(url):
    # headers = [
    #     {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
    #     {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 '
    #                    'Safari/535.11'},
    #     {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]

    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:62.0) Gecko/20100101 Firefox/62.0'}
    res = requests.get(url, headers=headers)
    res.encoding = 'utf8'  # 不加这个会乱码

    return res


# 提取学校各学院专业
def find_major(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')  # 参数2指定解析器
    l_department = []
    school = ''
    cnt_dp = 0  # 学校院系统计
    cnt_mj = 0  # 院系专业统计

    # 提取学校名称
    untreated_title = soup.find('div', class_='school')
    if untreated_title is not None:  # 有部分 404 页面
        school = untreated_title.text

    # # 提取院系名
    # ps = soup.select('div.pro_content_y p')  # 利用 CSS选择器 .select()
    # if ps is not None:
    #     for p in ps:
    #         department.append(p)
    #
    # # 提取专业
    # lis = soup.select('div.pro_content_y li')
    # if lis is not None:
    #     for li in lis:
    #         major.append(li.a.text)  # 提取专业名
    #         major_url.append(li.a.get('href'))  # 提取超链接

    # 提取院系专业
    # https://www.jianshu.com/p/74c1acd7ca8b
        for block in soup.find_all('div', class_='pro_content_y'):
            p = block.find('p')  # 院系名
            if p is not None:
                cnt_dp += 1
            majors = block.find_all('a')  # 对应专业
            l_major_name = []
            l_major_url = []
            if p is not None:  # none 没有.text(), 会报错
                for m in majors:
                    cnt_mj += 1
                    l_major_name.append(m.text)
                    l_major_url.append('https://souky.eol.cn' + m.get('href'))
                l_department.append([p.text, l_major_name, l_major_url, cnt_mj])
            cnt_mj = 0  # 清零统计

    # 整合数据
    if school != '':  # 排除 404 页面的空数据
        info = [school, l_department, cnt_dp]
        return info


# 导出数据
def data_export(data):
    # 导出院校基本信息表
    wb = openpyxl.Workbook()  # 坑爹, Workbook() 的 W 要大写
    wb.guess_types = True  # 自动匹配数据类型
    ws = wb.active
    mark1 = 1  # 第一列标记
    mark2 = 1  # 第二列标记
    mark3 = 1  # 第三第四列标记
    for school in range(len(data)):
        d1 = ws.cell(row=mark1, column=1)  # 校名列
        d1.value = data[school][0]
        for department in range(data[school][2]):  # 填充
            d2 = ws.cell(row=mark2, column=2)
            d2.value = data[school][1][department][0]
            for major in range(data[school][1][department][3]):
                d3 = ws.cell(row=mark3, column=3)  # 专业名列
                d4 = ws.cell(row=mark3, column=4)  # 专业链接列
                d3.value = data[school][1][department][1][major]
                d4.value = data[school][1][department][2][major]
                mark3 += 1
            for i in range(mark2, mark3):
                dt = ws.cell(row=i, column=2)
                dt.value = data[school][1][department][0]
            mark2 += data[school][1][department][3]
        for i in range(mark1, mark3):  # 填充
                dt = ws.cell(row=i, column=1)
                dt.value = data[school][0]
        mark1 = mark3  # 第一列标记移动到前一个学校最后一个专业后一格

    wb.save('专业信息.xlsx')


def main():
    school_basic_url = 'https://souky.eol.cn/HomePage/school_prolist_'
    total_info = []
    for i in range(2, 1246):  # 人工识别总页数范围
        single_info = find_major(get_url(school_basic_url + str(i) + '.html'))
        if single_info is not None:  # 排除 404 页面的空数据
            total_info.append(single_info)
    # print(total_info[0][1][0][3])
    data_export(total_info)


if __name__ == '__main__':
    main()
